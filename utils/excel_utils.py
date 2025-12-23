import openpyxl


def get_test_data(file_path,sheet_name="Sheet1"):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheet_name]
    data=[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        test_id,username,password,test_name,result=row
        data.append((test_id,username,password))
    return data

def update_result(file_path,test_id,result,sheet_name="Sheet1"):
    workbook=openpyxl.load_workbook(file_path)
    sheet=workbook[sheet_name]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == test_id:
            row[4].value = result
            break
    workbook.save(file_path)

